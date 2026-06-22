import io
import sqlite3
from datetime import datetime

from flask import Blueprint, flash, redirect, render_template, request, send_file, session
from database import conectar

estoque_bp = Blueprint('estoque', __name__)


def _formatar_moeda(valor):
    """Formata um número no padrão monetário brasileiro: R$ 1.234,56."""
    texto = f"{float(valor or 0):,.2f}"
    return "R$ " + texto.replace(",", "X").replace(".", ",").replace("X", ".")


def _data_hora(valor):
    """Separa 'YYYY-MM-DD HH:MM:SS' em (data dd/mm/aaaa, hora HH:MM)."""
    try:
        dt = datetime.strptime(valor, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M")
    except (ValueError, TypeError):
        return (valor or ""), ""


@estoque_bp.route('/estoque')
def listar_estoque():
    db = conectar()

    busca = request.args.get('busca', '')
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 5

    if pagina < 1:
        pagina = 1

    termo_busca = f'%{busca}%'
    parametros_busca = (termo_busca,) * 3

    total_filtrado = db.execute("""
        SELECT COUNT(*) AS total
        FROM pecas
        WHERE nome LIKE ?
           OR descricao LIKE ?
           OR fornecedor LIKE ?
    """, parametros_busca).fetchone()['total']

    total_paginas = max(1, (total_filtrado + por_pagina - 1) // por_pagina)

    if pagina > total_paginas:
        pagina = total_paginas

    offset = (pagina - 1) * por_pagina

    pecas = db.execute("""
        SELECT *
        FROM pecas
        WHERE nome LIKE ?
           OR descricao LIKE ?
           OR fornecedor LIKE ?
        ORDER BY nome
        LIMIT ? OFFSET ?
    """, parametros_busca + (por_pagina, offset)).fetchall()

    total_pecas = db.execute(
        "SELECT COUNT(*) AS total FROM pecas"
    ).fetchone()['total']

    estoque_baixo = db.execute(
        "SELECT COUNT(*) AS total FROM pecas WHERE quantidade <= estoque_minimo"
    ).fetchone()['total']

    total_movimentacoes = db.execute(
        "SELECT COUNT(*) AS total FROM movimentacoes_estoque"
    ).fetchone()['total']

    db.close()

    return render_template(
        'estoque.html',
        pecas=pecas,
        busca=busca,
        pagina=pagina,
        total_paginas=total_paginas,
        total_pecas=total_pecas,
        estoque_baixo=estoque_baixo,
        total_movimentacoes=total_movimentacoes
    )


@estoque_bp.route('/estoque/cadastrar', methods=['POST'])
def cadastrar_peca():
    nome = request.form['nome']
    descricao = request.form['descricao']
    fornecedor = request.form['fornecedor']

    try:
        quantidade = int(request.form['quantidade'] or 0)
        estoque_minimo = int(request.form['estoque_minimo'] or 0)
        preco_unitario = float(request.form['preco_unitario'] or 0)
    except ValueError:
        flash("Quantidade, estoque mínimo e preço devem ser números válidos.", "erro")
        return redirect('/estoque')

    if quantidade < 0 or estoque_minimo < 0 or preco_unitario < 0:
        flash("Quantidade, estoque mínimo e preço não podem ser negativos.", "erro")
        return redirect('/estoque')

    db = conectar()
    db.execute("""
        INSERT INTO pecas
        (nome, descricao, quantidade, estoque_minimo, preco_unitario, fornecedor)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (nome, descricao, quantidade, estoque_minimo, preco_unitario, fornecedor))
    db.commit()
    db.close()

    flash(f'Peça "{nome}" cadastrada com sucesso.', "sucesso")
    return redirect('/estoque')


@estoque_bp.route('/estoque/editar/<int:id>')
def editar_peca(id):
    db = conectar()
    peca = db.execute("SELECT * FROM pecas WHERE id = ?", (id,)).fetchone()
    db.close()

    if peca is None:
        flash("Peça não encontrada.", "erro")
        return redirect('/estoque')

    return render_template('editar_peca.html', peca=peca)


@estoque_bp.route('/estoque/atualizar/<int:id>', methods=['POST'])
def atualizar_peca(id):
    nome = request.form['nome']
    descricao = request.form['descricao']
    fornecedor = request.form['fornecedor']

    try:
        quantidade = int(request.form['quantidade'] or 0)
        estoque_minimo = int(request.form['estoque_minimo'] or 0)
        preco_unitario = float(request.form['preco_unitario'] or 0)
    except ValueError:
        flash("Quantidade, estoque mínimo e preço devem ser números válidos.", "erro")
        return redirect(f'/estoque/editar/{id}')

    if quantidade < 0 or estoque_minimo < 0 or preco_unitario < 0:
        flash("Quantidade, estoque mínimo e preço não podem ser negativos.", "erro")
        return redirect(f'/estoque/editar/{id}')

    db = conectar()
    db.execute("""
        UPDATE pecas
        SET nome = ?, descricao = ?, quantidade = ?, estoque_minimo = ?,
            preco_unitario = ?, fornecedor = ?
        WHERE id = ?
    """, (nome, descricao, quantidade, estoque_minimo, preco_unitario, fornecedor, id))
    db.commit()
    db.close()

    flash("Peça atualizada com sucesso.", "sucesso")
    return redirect('/estoque')


@estoque_bp.route('/estoque/excluir/<int:id>', methods=['POST'])
def excluir_peca(id):
    db = conectar()
    try:
        db.execute("DELETE FROM pecas WHERE id = ?", (id,))
        db.commit()
        flash("Peça excluída com sucesso.", "sucesso")
    except sqlite3.IntegrityError:
        db.rollback()
        flash(
            "Não é possível excluir esta peça porque ela possui movimentações ou "
            "está vinculada a ordens de serviço.",
            "erro",
        )
    finally:
        db.close()

    return redirect('/estoque')


@estoque_bp.route('/estoque/movimentar/<int:id>')
def movimentar_peca(id):
    db = conectar()
    peca = db.execute("SELECT * FROM pecas WHERE id = ?", (id,)).fetchone()
    db.close()

    if peca is None:
        flash("Peça não encontrada.", "erro")
        return redirect('/estoque')

    return render_template('movimentar_estoque.html', peca=peca)


@estoque_bp.route('/estoque/movimentar/<int:id>', methods=['POST'])
def salvar_movimentacao(id):
    tipo = request.form['tipo']
    observacao = request.form['observacao']

    if tipo not in ('Entrada', 'Saida'):
        flash("Tipo de movimentação inválido.", "erro")
        return redirect(f'/estoque/movimentar/{id}')

    try:
        quantidade = int(request.form['quantidade'] or 0)
    except ValueError:
        flash("A quantidade deve ser um número válido.", "erro")
        return redirect(f'/estoque/movimentar/{id}')

    if quantidade <= 0:
        flash("A quantidade deve ser maior que zero.", "erro")
        return redirect(f'/estoque/movimentar/{id}')

    db = conectar()
    peca = db.execute("SELECT quantidade FROM pecas WHERE id = ?", (id,)).fetchone()

    if peca is None:
        db.close()
        flash("Peça não encontrada.", "erro")
        return redirect('/estoque')

    estoque_atual = peca['quantidade']

    if tipo == 'Entrada':
        novo_estoque = estoque_atual + quantidade
    else:
        novo_estoque = estoque_atual - quantidade
        if novo_estoque < 0:
            db.close()
            flash(
                f"Estoque insuficiente: há apenas {estoque_atual} unidade(s) disponível(is).",
                "erro",
            )
            return redirect(f'/estoque/movimentar/{id}')

    db.execute("UPDATE pecas SET quantidade = ? WHERE id = ?", (novo_estoque, id))
    db.execute("""
        INSERT INTO movimentacoes_estoque
        (peca_id, tipo, quantidade, data_movimentacao, observacao)
        VALUES (?, ?, ?, ?, ?)
    """, (
        id,
        tipo,
        quantidade,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        observacao,
    ))
    db.commit()
    db.close()

    flash(
        f"{tipo} de {quantidade} unidade(s) registrada. Novo estoque: {novo_estoque}.",
        "sucesso",
    )
    return redirect('/estoque')


@estoque_bp.route('/estoque/historico')
def historico_movimentacoes():
    db = conectar()
    movimentacoes = db.execute("""
        SELECT
            m.id,
            p.nome AS nome_peca,
            m.tipo,
            m.quantidade,
            m.data_movimentacao,
            m.observacao
        FROM movimentacoes_estoque m
        INNER JOIN pecas p ON p.id = m.peca_id
        ORDER BY m.data_movimentacao DESC
    """).fetchall()
    db.close()

    return render_template('historico_movimentacoes.html', movimentacoes=movimentacoes)


def _carregar_dados_relatorio():
    db = conectar()
    pecas = db.execute("SELECT * FROM pecas ORDER BY nome").fetchall()
    movimentacoes = db.execute("""
        SELECT m.id, p.nome AS nome_peca, m.tipo, m.quantidade, m.data_movimentacao
        FROM movimentacoes_estoque m
        INNER JOIN pecas p ON p.id = m.peca_id
        ORDER BY m.data_movimentacao DESC
    """).fetchall()
    db.close()
    return pecas, movimentacoes


@estoque_bp.route('/estoque/relatorio/pdf')
def gerar_relatorio_pdf():
    import tempfile
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfgen import canvas
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    pecas, movimentacoes = _carregar_dados_relatorio()

    total_pecas = len(pecas)
    qtd_total = sum(p['quantidade'] for p in pecas)
    itens_baixo = sum(1 for p in pecas if p['quantidade'] <= p['estoque_minimo'])
    total_mov = len(movimentacoes)
    valor_total = sum(p['quantidade'] * (p['preco_unitario'] or 0) for p in pecas)

    responsavel = session.get("funcionario_nome") or "—"
    emissao = datetime.now().strftime("%d/%m/%Y %H:%M")

    azul = colors.HexColor("#123a66")
    azul_claro = colors.HexColor("#eef2ff")
    cinza = colors.HexColor("#cccccc")
    verde_soft = colors.HexColor("#dcfce7")
    verde = colors.HexColor("#166534")
    verm_soft = colors.HexColor("#fee2e2")
    verm = colors.HexColor("#991b1b")

    L = A4[0] - 24 * mm

    est_titulo = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=13, leading=15, textColor=azul)
    est_head = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, leading=11, alignment=2)
    est_band = ParagraphStyle("b", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=colors.white)
    est_cell = ParagraphStyle("c", fontName="Helvetica", fontSize=7.5, leading=9)
    est_cellb = ParagraphStyle("cb", fontName="Helvetica-Bold", fontSize=7.5, leading=9)
    est_th = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=colors.white)
    est_kpi = ParagraphStyle("k", fontName="Helvetica-Bold", fontSize=12, leading=14, alignment=1, textColor=azul)

    def banda(texto):
        t = Table([[Paragraph(texto, est_band)]], colWidths=[L])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    elementos = []

    # Cabeçalho corporativo (mesma identidade do PDF da OS)
    cab = Table([[
        Paragraph("SISTEMA DE ASSISTÊNCIA TÉCNICA"
                  "<br/><font size=9 color='#555555'>Relatório Geral de Estoque</font>", est_titulo),
        Paragraph(f"Emitido em:<br/>{emissao}<br/><font size=8>Responsável: {responsavel}</font>", est_head),
    ]], colWidths=[L * 0.6, L * 0.4])
    cab.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 1, azul),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(cab)
    elementos.append(Spacer(1, 10))

    # Resumo executivo
    elementos.append(banda("RESUMO EXECUTIVO"))
    elementos.append(Spacer(1, 6))

    def kpi(valor, rotulo):
        return Paragraph(f"{valor}<br/><font size=6.5 color='#555555'>{rotulo}</font>", est_kpi)

    resumo = Table([[
        kpi(total_pecas, "Peças cadastradas"),
        kpi(qtd_total, "Qtd. em estoque"),
        kpi(itens_baixo, "Estoque baixo"),
        kpi(total_mov, "Movimentações"),
        kpi(_formatar_moeda(valor_total), "Valor do estoque"),
    ]], colWidths=[L / 5.0] * 5)
    resumo.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, cinza),
        ("BACKGROUND", (0, 0), (-1, -1), azul_claro),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(resumo)
    elementos.append(Spacer(1, 12))

    # Tabela de peças
    elementos.append(banda("TABELA DE PEÇAS"))
    elementos.append(Spacer(1, 6))

    cabecalhos = ["ID", "Nome", "Descrição", "Qtd", "Est. Mín", "Preço Unit.", "Fornecedor", "Status"]
    dados = [[Paragraph(h, est_th) for h in cabecalhos]]
    estilo_pecas = [
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("GRID", (0, 0), (-1, -1), 0.5, cinza),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (3, 1), (5, -1), "CENTER"),
    ]
    if not pecas:
        dados.append([Paragraph("Nenhuma peça cadastrada.", est_cell), "", "", "", "", "", "", ""])
        estilo_pecas.append(("SPAN", (0, 1), (-1, 1)))
    for i, p in enumerate(pecas, start=1):
        baixo = p['quantidade'] <= p['estoque_minimo']
        dados.append([
            Paragraph(str(p['id']), est_cell),
            Paragraph(p['nome'] or "", est_cell),
            Paragraph(p['descricao'] or "", est_cell),
            Paragraph(str(p['quantidade']), est_cell),
            Paragraph(str(p['estoque_minimo']), est_cell),
            Paragraph(_formatar_moeda(p['preco_unitario']), est_cell),
            Paragraph(p['fornecedor'] or "", est_cell),
            Paragraph("Estoque Baixo" if baixo else "Normal", est_cellb),
        ])
        estilo_pecas.append(("BACKGROUND", (7, i), (7, i), verm_soft if baixo else verde_soft))
        estilo_pecas.append(("TEXTCOLOR", (7, i), (7, i), verm if baixo else verde))

    tabela = Table(dados, colWidths=[24, 88, 108, 34, 42, 62, 82, 72], repeatRows=1)
    tabela.setStyle(TableStyle(estilo_pecas))
    elementos.append(tabela)
    elementos.append(Spacer(1, 14))

    # Histórico de movimentações
    elementos.append(banda("HISTÓRICO DE MOVIMENTAÇÕES"))
    elementos.append(Spacer(1, 6))

    cabecalhos2 = ["Peça", "Tipo", "Quantidade", "Data", "Hora"]
    dados2 = [[Paragraph(h, est_th) for h in cabecalhos2]]
    estilo_mov = [
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("GRID", (0, 0), (-1, -1), 0.5, cinza),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]
    if not movimentacoes:
        dados2.append([Paragraph("Nenhuma movimentação registrada.", est_cell), "", "", "", ""])
        estilo_mov.append(("SPAN", (0, 1), (-1, 1)))
    for m in movimentacoes:
        data, hora = _data_hora(m['data_movimentacao'])
        dados2.append([
            Paragraph(m['nome_peca'] or "", est_cell),
            Paragraph(m['tipo'] or "", est_cell),
            Paragraph(str(m['quantidade']), est_cell),
            Paragraph(data, est_cell),
            Paragraph(hora, est_cell),
        ])

    tabela2 = Table(dados2, colWidths=[L * 0.40, L * 0.15, L * 0.17, L * 0.16, L * 0.12], repeatRows=1)
    tabela2.setStyle(TableStyle(estilo_mov))
    elementos.append(tabela2)

    # Canvas com rodapé e paginação "Página X de Y"
    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            canvas.Canvas.__init__(self, *args, **kwargs)
            self._paginas_salvas = []

        def showPage(self):
            self._paginas_salvas.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._paginas_salvas)
            for estado in self._paginas_salvas:
                self.__dict__.update(estado)
                self._rodape(total)
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

        def _rodape(self, total):
            self.setFont("Helvetica", 7)
            self.setFillColor(colors.HexColor("#888888"))
            self.drawString(12 * mm, 8 * mm,
                            "Relatório gerado automaticamente pelo Sistema de Assistência Técnica.")
            self.drawRightString(A4[0] - 12 * mm, 8 * mm,
                                 f"Página {self._pageNumber} de {total}")

    arquivo_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf = SimpleDocTemplate(
        arquivo_temp.name, pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=16 * mm,
        title="Relatório de Estoque",
    )
    pdf.build(elementos, canvasmaker=NumberedCanvas)

    return send_file(
        arquivo_temp.name,
        as_attachment=True,
        download_name="relatorio_estoque.pdf",
    )


@estoque_bp.route('/estoque/relatorio/excel')
def gerar_relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    pecas, movimentacoes = _carregar_dados_relatorio()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="123A66")
    center = Alignment(horizontal="center", vertical="center")
    baixo_fill = PatternFill("solid", fgColor="FEE2E2")
    baixo_font = Font(color="991B1B", bold=True)
    normal_font = Font(color="166534")
    thin = Side(style="thin", color="D9D9D9")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def estilizar_cabecalho(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = borda
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def autoajustar(ws):
        for col in ws.columns:
            largura = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(largura + 3, 55)

    wb = Workbook()

    # Aba 1 — Estoque
    ws = wb.active
    ws.title = "Estoque"
    ws.append(["ID", "Nome", "Descrição", "Quantidade", "Estoque Mínimo",
               "Preço Unitário", "Fornecedor", "Status"])
    for p in pecas:
        baixo = p['quantidade'] <= p['estoque_minimo']
        ws.append([
            p['id'], p['nome'], p['descricao'], p['quantidade'], p['estoque_minimo'],
            float(p['preco_unitario'] or 0), p['fornecedor'],
            "Estoque Baixo" if baixo else "Normal",
        ])
        linha = ws.max_row
        ws.cell(row=linha, column=6).number_format = 'R$ #,##0.00'
        if baixo:
            for c in range(1, 9):
                ws.cell(row=linha, column=c).fill = baixo_fill
            ws.cell(row=linha, column=8).font = baixo_font
        else:
            ws.cell(row=linha, column=8).font = normal_font
    estilizar_cabecalho(ws, 8)
    autoajustar(ws)

    # Aba 2 — Movimentações
    ws2 = wb.create_sheet("Movimentações")
    ws2.append(["ID da Movimentação", "Peça", "Tipo", "Quantidade", "Data", "Hora"])
    for m in movimentacoes:
        data, hora = _data_hora(m['data_movimentacao'])
        ws2.append([m['id'], m['nome_peca'], m['tipo'], m['quantidade'], data, hora])
    estilizar_cabecalho(ws2, 6)
    autoajustar(ws2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="estoque.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
